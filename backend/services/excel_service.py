from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from typing import List
from models.schemas import TestCase

class ExcelService:
    @staticmethod
    def export_test_cases(test_cases: List[TestCase]) -> BytesIO:
        """Export test cases to Excel file"""
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Define headers
        headers = [
            'Test Case ID', 'Title', 'Steps', 'Expected Result', 
            'Priority', 'Status', 'Created Date', 'Updated Date'
        ]
        
        # Write headers with styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write data
        for row_num, test_case in enumerate(test_cases, 2):
            ws.cell(row=row_num, column=1, value=f"TC-{test_case.id:04d}")
            ws.cell(row=row_num, column=2, value=test_case.title)
            ws.cell(row=row_num, column=3, value=test_case.steps.replace('\n', ' | '))
            ws.cell(row=row_num, column=4, value=test_case.expected_result.replace('\n', ' | '))
            ws.cell(row=row_num, column=5, value=test_case.priority.capitalize())
            ws.cell(row=row_num, column=6, value=test_case.status.capitalize())
            ws.cell(row=row_num, column=7, value=test_case.created_at.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=8, value=test_case.updated_at.strftime('%Y-%m-%d'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output